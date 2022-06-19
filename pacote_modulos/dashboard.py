from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.line import LineProperties
from datetime import datetime, timedelta
import qrcode


def inicializa_planilha():
    """
    Cria a planilha da carteira.\n
    :return: variável de referência da planilha, variável de referência de folha da planilha
    """
    _planilha = Workbook()  # cria planilha
    _folha = _planilha.active  # acessa folha da planilha
    return _planilha, _folha


def formatacao_inicial(_folha):
    """
    Define aspectos visuais básicos da planilha.\n
    :param _folha: variável de referência de folha da planilha
    :return:
    """
    _folha.title = "Dashboard"  # altera nome da folha

    for coluna in "ABCDEFGH":  # ajuste de largura de coluna
        _folha.column_dimensions[coluna].width = 27

    for linha in range(1, 3):  # ajuste de altura de linha
        _folha.row_dimensions[linha].height = 27


def criar_titulo(_folha, ini_row, ini_col, fim_row, fim_col, nome, tamanho="12"):
    """
    Cria célula de título.\n
    :param _folha: variável de referência de folha da planilha
    :param ini_row: linha inicial
    :param ini_col: coluna inicial
    :param fim_row: linha final
    :param fim_col: coluna final
    :param nome: título
    :param tamanho: tamanho da fonte
    :return:
    """
    celula = _folha.merge_cells(start_row=ini_row, start_column=ini_col, end_row=fim_row, end_column=fim_col)
    celula = _folha.cell(row=ini_row, column=ini_col, value=nome)
    celula.alignment = Alignment(horizontal="center", vertical="center")
    celula.font = Font(bold=True, size=tamanho)


def celulas_fixas(_folha):
    """
    Cria as células que têm posições fixas na planilha.\n
    :param _folha: variável de referência de folha da planilha
    :return:
    """

    # Criação de título
    criar_titulo(_folha, 1, 1, 2, 8, "Resumo da Carteira", "20")

    # Espaço para ações
    criar_titulo(_folha, 3, 1, 4, 4, "Ações", "16")
    criar_titulo(_folha, 5, 1, 5, 1, "Nome")
    criar_titulo(_folha, 5, 2, 5, 2, "Quantidade")
    criar_titulo(_folha, 5, 3, 5, 3, "Valor da ação (R$)")
    criar_titulo(_folha, 5, 4, 5, 4, "Valor acumulado (R$)")

    # Espaço para moedas
    criar_titulo(_folha, 3, 5, 4, 8, "Moedas", "16")
    criar_titulo(_folha, 5, 5, 5, 5, "Nome")
    criar_titulo(_folha, 5, 6, 5, 6, "Quantidade")
    criar_titulo(_folha, 5, 7, 5, 7, "Valor da ação (R$)")
    criar_titulo(_folha, 5, 8, 5, 8, "Valor acumulado (R$)")


def listar_acoes(_carteira):
    """
    Extrai as ações da carteira.\n
    :param _carteira: variável dicionário que consolida informações da carteira
    :return: lista de ações da carteira, com nome, quantidade, e valor atual
    """
    acoes = []
    for ativo in _carteira["acao"]:
        acoes.append([ativo["Nome"], ativo["Quantidade"], ativo["preco_atualizado"]])

    return acoes


def apresentar_acoes(_folha, _acoes):
    """
    Lista as ações da carteira na planilha.\n
    :param _folha: variável de referência de folha da planilha
    :param _acoes: lista de ações da carteira
    :return:
    """
    num_acoes = len(_acoes)

    for i in range(num_acoes):  # Iterador das ações
        _folha.cell(row=(6 + i), column=1, value=_acoes[i][0])  # Nome da ação
        _folha.cell(row=(6 + i), column=2, value=float(_acoes[i][1]))  # Quantidade na carteira
        cotacao = _folha.cell(row=(6 + i), column=3, value=float(_acoes[i][2]))  # Valor atual da ação
        cotacao.number_format = "R$#,##0.00"  # Formato de moeda real
        contrib = _folha.cell(row=(6 + i), column=4,
                              value="=B" + str(6 + i) + "*C" + str(6 + i))  # Contribuição da ação (valor*quantidade)
        contrib.number_format = "R$#,##0.00"  # Formato de moeda real


def listar_moedas(_carteira):
    """
    Extrai as moedas da carteira.\n
    :param _carteira: variável dicionário que consolida informações da carteira
    :return: lista de moedass da carteira, com nome, quantidade, e valor atual
    """
    moedas = []
    for ativo in _carteira["moeda"]:
        moedas.append([ativo["Nome"], ativo["Quantidade"], ativo["preco_atualizado"]])

    return moedas


def apresentar_moedas(_folha, _moedas):
    """
    Lista as moedas da carteira na planilha.\n
    :param _folha: variável de referência de folha da planilha
    :param _moedas: lista de moedass da carteira
    :return:
    """
    num_moedas = len(_moedas)

    # Apresentação de Moedas
    for i in range(num_moedas):  # Iterador de moedas
        _folha.cell(row=(6 + i), column=5, value=_moedas[i][0])  # Nome da moeda
        _folha.cell(row=(6 + i), column=6, value=float(_moedas[i][1]))  # Quantidade na carteira
        cotacao = _folha.cell(row=(6 + i), column=7, value=float(_moedas[i][2]))  # Valor atual da moeda
        cotacao.number_format = "R$#,##0.00"  # Formato de moeda real
        contrib = _folha.cell(row=(6 + i), column=8,
                              value="=F" + str(6 + i) + "*G" + str(6 + i))  # Contribuição da moeda (valor*quantidade)
        contrib.number_format = "R$#,##0.00"  # Formato de moeda real


def qtd_linhas(_acoes, _moedas):
    """
    Retorna a quantidade variável de linhas.\n
    :param _acoes: lista de ações da carteira
    :param _moedas: lista de moedas da carteira
    :return: máximo dentre as duas quantidades, de ações ou de moedas
    """
    num_acoes = len(_acoes)
    num_moedas = len(_moedas)

    return max(num_moedas, num_acoes)


def totais_acoes(_folha, _num_linhas):
    """
    Adiciona informações do total das ações da carteira.\n
    :param _folha: variável de referência de folha da planilha
    :param _num_linhas: quantidade variável de linhas
    :return:
    """
    criar_titulo(_folha, 6 + _num_linhas, 1, 6 + _num_linhas, 1, "Total Ações")
    _folha.cell(row=(6 + _num_linhas), column=2,
                value="=SUM(B6:B" + str(5 + _num_linhas) + ")")  # Soma das quantidades de ações
    soma_val = _folha.cell(row=(6 + _num_linhas), column=3,
                           value="=SUM(C6:C" + str(5 + _num_linhas) + ")")  # Soma dos valores das ações
    soma_val.number_format = "R$#,##0.00"  # Formato de moeda real
    soma_contrib = _folha.cell(row=(6 + _num_linhas), column=4,
                               value="=SUM(D6:D" + str(5 + _num_linhas) + ")")  # Soma das contribuições das ações
    soma_contrib.number_format = "R$#,##0.00"  # Formato de moeda real


def totais_moedas(_folha, _num_linhas):
    """
    Adiciona informações do total das moedas da carteira.\n
    :param _folha: variável de referência de folha da planilha
    :param _num_linhas: quantidade variável de linhas
    :return:
    """
    criar_titulo(_folha, 6 + _num_linhas, 5, 6 + _num_linhas, 5, "Total Moedas")
    _folha.cell(row=(6 + _num_linhas), column=6,
                value="=SUM(F6:F" + str(5 + _num_linhas) + ")")  # Soma das quantidades de moedas
    soma_val = _folha.cell(row=(6 + _num_linhas), column=7,
                           value="=SUM(G6:G" + str(5 + _num_linhas) + ")")  # Soma dos valores das ações
    soma_val.number_format = "R$#,##0.00"  # Formato de moeda real
    soma_contrib = _folha.cell(row=(6 + _num_linhas), column=8,
                               value="=SUM(H6:H" + str(5 + _num_linhas) + ")")  # Soma das contribuições das moedas
    soma_contrib.number_format = "R$#,##0.00"  # Formato de moeda real


def total_carteira(_folha, _num_linhas):
    """
    Apresentação do Total da carteira.\n
    :param _folha: variável de referência de folha da planilha
    :param _num_linhas: quantidade variável de linhas
    :return:
    """
    criar_titulo(_folha, 9 + _num_linhas, 4, 10 + _num_linhas, 5, "Valor da Carteira", "16")
    criar_titulo(_folha, 11 + _num_linhas, 4, 11 + _num_linhas, 4, "Quantidade")
    criar_titulo(_folha, 11 + _num_linhas, 5, 11 + _num_linhas, 5, "Valor acumulado total (R$)")
    soma_val = _folha.cell(row=12 + _num_linhas, column=4, value="=B" + str(6 + _num_linhas) + "+F"
                           + str(6 + _num_linhas))  # Soma das quantidades de ações e de moedas
    soma_val.number_format = "R$#,##0.00"  # Formato de moeda real
    soma_contrib = _folha.cell(row=12 + _num_linhas, column=5, value="=D" + str(6 + _num_linhas) + "+H"
                               + str(6 + _num_linhas))  # Soma dos valores acumulados de ações e de moedas
    soma_contrib.number_format = "R$#,##0.00"  # Formato de moeda real


def cria_hist(_planilha, _carteira):
    """
    Cria folha de histórico dos valores dos ativos da carteira.\n
    :param _planilha: variável de referência de folha da planilha
    :param _carteira: variável dicionário que consolida informações da carteira
    :return: última linha com dados de valores da ação que mais vale atualmente
    """
    _folha = _planilha.create_sheet("Histórico")  # Cria uma folha para histórico dos valores de ações

    dados_acao = _carteira["acao"]  # Resgata dados sobre ações
    dados_acao = sorted(dados_acao, key=lambda x: -x['preco_atualizado'])  # Ordena os dados em ordem decresc. de valor
    num_acoes = len(dados_acao)  # Qtd de ações
    _ultima_linha = 0  # valor temporário da última linha com dados de valores da ação que mais vale atualmente

    for i in range(num_acoes):  # iteração de ações
        criar_titulo(_folha, 1, 2*i+1, 1, 2*i+2, dados_acao[i]["Nome"])  # Nome da ação
        _folha.cell(row=2, column=2*i+1, value="Data")  # Cabeçalho
        _folha.cell(row=2, column=2*i+2, value="Close")  # Cabeçalho

        dados_historicos = dados_acao[i]["preco_historico"]  # Histórico de ações desta ação

        valores = dados_historicos.values.tolist()
        valores = [valor[0] for valor in valores]  # lista de valores desta ação
        datas = dados_historicos.index.tolist()  # lista de datas desta ação
        datas = [datetime.strptime(str(date), "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d") for date in datas]  # -> string

        _qtd_linhas = len(datas)  # qtd de datas

        if i == 0:  # verifica se a ação atual é a mais valiosa
            _ultima_linha = 2 + _qtd_linhas  # atualiza _ultima_linha

        for j in range(_qtd_linhas):  # itera sobre as datas
            _folha.cell(row=3+j, column=2*i+1, value=datas[j])  # data
            _cell = _folha.cell(row=3+j, column=2*i+2, value=valores[j])  # valor na data
            _cell.number_format = "R$#,##0.00"  # formatação

    dados_moedas = _carteira["moeda"]  # Resgata dados sobre moedas
    num_moedas = len(dados_moedas)  # qtd de moedas

    for i in range(num_moedas):  # iteração de moedas
        criar_titulo(_folha, 1, 2*num_acoes+2 * i + 1, 1, 2*num_acoes+2 * i + 2, dados_moedas[i]["Nome"])
        # nome da moeda
        _folha.cell(row=2, column=2*num_acoes+2 * i + 1, value="Data")  # cabeçalho
        _folha.cell(row=2, column=2*num_acoes+2 * i + 2, value="Close")  # cabeçalho

        dados_historicos = dados_moedas[i]["preco_historico"]  # Histórico de valores desta moeda

        valores = dados_historicos.values.tolist()
        valores = [valor[0] for valor in valores]  # lista de valores desta moeda
        datas = dados_historicos.index.tolist()  # lista de datas desta moeda
        datas = [datetime.strptime(str(date), "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d") for date in datas]

        _qtd_linhas = len(datas)  # qtd de datas

        for j in range(_qtd_linhas):  # itera sobre datas
            _cell = _folha.cell(row=3 + j, column=2*num_acoes+2 * i + 1, value=datas[j])  # data
            _folha.cell(row=3 + j, column=2*num_acoes+2 * i + 2, value=valores[j])  # valor na data
            _cell.number_format = "R$#,##0.00"  # formatação
    
    return _ultima_linha


def graf_barras1(_folha, _num_linhas):  # Criação e apresentação do primeiro gráfico
    """_ Cria um gráfico de barras que analisa a composição da carteira por ação_

    Args:
        _folha: _Acessa os dados da folha da planilha_
        _num_linhas: _Acessa a última linha de referência da tabela do dashboard_
    """
    graf_1 = BarChart()  # Gráfico de barras
    graf_1.type = "col"  # Tipo de gráfico
    graf_1.style = 10  # Tamanho do gráfico
    graf_1.title = "Composição da carteira (por ação)"  # Título do gráfico
    graf_1.y_axis.title = 'Valor de cada ação'  # Título do eixo y
    graf_1.x_axis.title = 'Ações'  # Título do eixo x
    graf_1.legend = None  # Exclui legenda

    data = Reference(_folha, min_col=4, min_row=6, max_row=_num_linhas+5)  # Seleciona o valor acumulado de cada ação
    cats = Reference(_folha, min_col=1, min_row=6, max_row=_num_linhas+5)  # Seleciona o nome de cada ação
    graf_1.add_data(data, titles_from_data=False)  # Adiciona o valor acumulado de cada ação
    graf_1.set_categories(cats)  # Adiciona o nome de cada ação
    graf_1.shape = 4  # Formato do gráfico

    _folha.add_chart(graf_1, "A20")  # Adiciona o gráfico na planilha


def graf_barras2(_folha, _num_linhas):  # Criação e apresentação do segundo gráfico
    """_Cria um gráfico de barras que analisa a composição da carteira por moeda_

    Args:
        _folha: _Acessa os dados da folha da planilha_
        _num_linhas: _Acessa a última linha de referência da tabela do dashboard_
    """
    graf_2 = BarChart()  # Gráfico de barras
    graf_2.type = "col"  # Tipo de gráfico
    graf_2.style = 10  # Tamanho do gráfico
    graf_2.title = "Composição da carteira (por moeda)"  # Título do gráfico
    graf_2.y_axis.title = 'Valor de cada moeda'  # Título do eixo y
    graf_2.x_axis.title = 'Moedas'  # Título do eixo x
    graf_2.legend = None  # Exclui legenda

    data = Reference(_folha, min_col=8, min_row=6, max_row=_num_linhas+5)  # Seleciona o valor acumulado de cada moeda
    cats = Reference(_folha, min_col=5, min_row=6, max_row=_num_linhas+5)  # Seleciona o nome de cada moeda
    graf_2.add_data(data, titles_from_data=False)  # Adiciona o valor acumulado de cada moeda
    graf_2.set_categories(cats)  # Adiciona o nome de cada moeda
    graf_2.shape = 4  # Formato do gráfico

    _folha.add_chart(graf_2, "E20")  # Adiciona o gráfico na planilha


def graf_linhas3(_planilha, _ultima_linha):  # Criação e apresentação do terceiro gráfico
    """_Cria um gráfico de linhas que analisa o histórico da ação de maior valor na carteira_

    Args:
        _planilha: _Acessa os dados da folha da planilha_
        _ultima_linha: _Acessa a última linha de referência da tabela do histórico_
    """
    _folha = _planilha["Histórico"]  # Transfere para a worksheet selecionada

    graf_3 = LineChart()  # Gráfico de linhas
    graf_3.title = "Histórico da ação que mais vale na carteira (" + _folha["A1"].value + ")"  # Título do gráfico,
    # com nome da ação
    graf_3.style = 12  # Tamanho do gráfico
    graf_3.y_axis.title = "Valor da ação (em R$)"  # Título do eixo y
    graf_3.x_axis.number_format = "%Y-%m-%d"  # formato da data
    graf_3.x_axis.majorTimeUnit = "years"  # Unidade de tempo para o eixo x
    graf_3.x_axis.title = "Data"  # Título do eixo x
    graf_3.legend = None  # Exclui legenda

    dados = Reference(_folha, min_col=2, min_row=3, max_col=2, max_row=_ultima_linha)  # Seleciona o valor de
    # fechamento do ativo
    tempo = Reference(_folha, min_col=1, min_row=3, max_col=1, max_row=_ultima_linha)  # Seleciona a data de
    # fechamento do ativo
    graf_3.add_data(dados, titles_from_data=True)  # Adiciona o valor de fechamento do ativo
    graf_3.set_categories(tempo)  # Adiciona a data de fechamento do ativo

    # Estilização
    s1 = graf_3.series[0]
    s1.graphicalProperties.line.solidFill = "0000FF"  # Cor da linha
    s1.graphicalProperties.line.width = 25000  # Largura da linha

    _folha = _planilha["Dashboard"]  # Retorna para a worksheet selecionada

    _folha.add_chart(graf_3, "A35")  # Adiciona o gráfico na planilha


def salvar_excel(_planilha, nome_arquivo):
    """
    Salva planilha no diretório do usuário.\n
    :param _planilha: variável de referência da planilha
    :param nome_arquivo: nome atribuído ao arquivo
    :return:
    """
    _planilha.save(nome_arquivo + ".xlsx")


def criar_img(_valor, _nome):
    """
    Cria a imagem do QR Code.\n
    :param _valor: valor total da carteira
    :param _nome: nome da planilha associada
    :return:
    """
    data = f'O valor da carteira é R${_valor}'
    img = qrcode.make(data)

    img.save('qrcode_'+_nome+'.png')


def adc_img(_folha, _arquivo):
    """
    Adiciona a imagem criada em criar_img().\n
    :param _folha: variável de referência de folha da planilha
    :param _arquivo: nome do arquivo de imagem do QR Code
    :return:
    """
    imagem = Image(_arquivo)
    _folha.add_image(imagem, "E35")


def dashboard(_carteira, _nome):
    """
    Consolidação do módulo; cria dashboard com dados da carteira.\n
    :param _carteira: variável dicionário que consolida informações da carteira
    :param _nome: nome atribuído ao arquivo
    :return:
    """
    planilha, folha = inicializa_planilha()

    formatacao_inicial(folha)

    celulas_fixas(folha)

    acoes = listar_acoes(_carteira)
    moedas = listar_moedas(_carteira)

    apresentar_acoes(folha, acoes)
    apresentar_moedas(folha, moedas)

    num_linhas = qtd_linhas(acoes, moedas)

    totais_acoes(folha, num_linhas)
    totais_moedas(folha, num_linhas)

    total_carteira(folha, num_linhas)

    graf_barras1(folha, num_linhas)

    graf_barras2(folha, num_linhas)

    graf_linhas3(planilha, cria_hist(planilha, _carteira))

    criar_img(folha.cell(row=12 + num_linhas, column=5).value, _nome)

    adc_img(folha, 'qrcode_'+_nome+'.png')

    salvar_excel(planilha, _nome)
